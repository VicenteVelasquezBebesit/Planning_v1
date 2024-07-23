import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox
from datetime import datetime, timedelta
import pandas as pd
import locale
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email import encoders
import matplotlib.pyplot as plt
from openpyxl.styles import PatternFill
import os
import ast 

#Configuracion para ejecucion en Chile
locale.setlocale(locale.LC_TIME, 'es_CL.UTF-8')
#configuracion para envio de email
smtp_server = 'smtp.gmail.com'
smtp_port = 587
#configuracion de rutas
ruta_icono = r'C:\Users\HP\Documents\Codigos\Automatizaciones\config\logo.ico'
directorio = r'C:\Users\HP\Documents\Codigos\Automatizaciones\data'
archive_to = r'C:\Users\HP\Documents\Codigos\Automatizaciones\config\config.Totest.txt'
archive_cc = r'C:\Users\HP\Documents\Codigos\Automatizaciones\config\config.Cctest.txt'
ruta_RM = r'C:\Users\HP\Documents\Codigos\Automatizaciones\data\Planning_RM'
ruta_RG = r'C:\Users\HP\Documents\Codigos\Automatizaciones\data\Planning_Regiones'

def es_fin_de_semana(fecha):
    return fecha.weekday() >= 5  # 5 = Sábado, 6 = Domingo

def siguiente_dia_habil(fecha):
    while es_fin_de_semana(fecha):
        fecha += timedelta(days=1)
    return fecha

def formatearFecha():
    fecha_actual = datetime.now()
    
    manana = siguiente_dia_habil(fecha_actual + timedelta(days=1))
    pasado_manana = siguiente_dia_habil(manana + timedelta(days=1))
    
    manana_str = manana.strftime('%a').upper().rstrip('.')
    pasado_manana_str = pasado_manana.strftime('%a').upper().rstrip('.')
    if manana_str == 'MIÃ©':
        manana_str = 'MIÉ'
    elif pasado_manana_str == 'MIÃ©':
        pasado_manana_str = 'MIÉ'
    formato = f"{manana_str}/{pasado_manana_str}"
    return formato

def fotoDinamica(df, region):
# Ajuste de Mto Saldo a valores sin decimales    
    df['Mto Saldo'] = df['Mto Saldo'].map('{:.0f}'.format)

# Ajuste de Peso_caj a valores con 3 decimales
    df['Peso_Caj'] = df['Peso_Caj'].map('{:.3f}'.format)
    
# Se definen las propiedades del imagen de salida.  
    num_filas, num_cols = df.shape
    alto_tabla = num_filas * 0.4
    ancho_tabla = num_cols * 1.8 
    fig, ax = plt.subplots(figsize=(ancho_tabla, alto_tabla))
    ax.axis('tight')
    ax.axis('off')
    tabla = ax.table(cellText=df.values, colLabels= df.columns, loc='center')
    tabla.auto_set_font_size(False)
    tabla.set_fontsize(10)
    tabla.auto_set_column_width(col=list(range(len(df.columns))))
    tabla.scale(1, 1.5)
    
# Se consideran el tipo de Planning para realizar la colorizacion de acuerdo alos nombres disponibles en las columnas. RM = Region metropolitana, RG = regiones
    if region == 'RM':
        for i, valor in enumerate(df['N° OC']):
            if valor == 'Total':
                for key, cell in tabla.get_celld().items():
                    if key[0] > 0 and df.iloc[key[0] - 1]['N° OC'] == 'Total':
                        cell.set_facecolor('yellow')
    if region == "RG":
        region = 'Regiones'
        for i, valor in enumerate(df['N° Pedido']):
            if valor == 'Total':
                for key, cell in tabla.get_celld().items():
                    if key[0] > 0 and df.iloc[key[0] - 1]['N° Pedido'] == 'Total':
                        cell.set_facecolor('yellow')
    
# Se definen parametros de salida y se guarda la imagen
    fecha_actual = datetime.now()
    fecha_actual_str = fecha_actual.strftime('%d-%m')
    nombre_archivo = f'Planning_{region} {fecha_actual_str}.png'
    ruta = os.path.join(directorio, nombre_archivo)
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    plt.savefig(ruta, bbox_inches='tight', pad_inches=0.1)
    plt.close()

def casoRM(archivo_df):
    # Se recibe y se elimminan los datos que no se considentan
        df_rm = archivo_df
        df_rm = df_rm[~((df_rm['Vendedor']>1) & (df_rm["Vendedor"] <30))]
        columns_to_drop = ['Sucursal', 'Nombre Sucursal', 'Observacion', 'Vendedor', 'Atencion', 'N° Factura', 'Fec-A-Partir-De', 'Fec-Imp.', 'Fec-Desp.', 'Ubicacion', 'Precio_Lp', 'Precio_Bol', 'Dcto', 'Q-Ped', 'Q-Desp', 'Mto Pedido', 'Mto Desp', 'Largo_Caj', 'Ancho_Caj', 'Alto_Caj', 'Stk_Otros', 'Stk_Lib', '1A1-STOCK', '1J1-JAULA', '1P1-PISO', '1P2-PISO', '3I3-RACK', '3I7-RACK', '6I2-RACK3', 'Comentarios Articulo']
        columns_to_drop = [col for col in columns_to_drop if col in df_rm.columns]
        df_rm = df_rm.drop(columns=columns_to_drop, axis=1)
        
    # Se hacen cambios en el formato del excel recibido 
        df_rm['Va'] = 'Si'
        df_rm['Despacho'] = formatearFecha()
        filtro_vev = df_rm['F12'].str.contains('vev', case=False, na=False)
        df_rm.loc[filtro_vev, 'Cliente'] = df_rm.loc[filtro_vev, 'Cliente'] + ' VEV'
        df_rm = df_rm[~((df_rm['Q-Saldo']==0))]
        df_rm = df_rm.rename(columns={'Q-Saldo': 'Pedido'})
        df_rm = df_rm.rename(columns={'Embaje': 'EMB'})
        
    # Se realizan calculos de volumen
        df_rm['Peso_Caj'] = df_rm['Vol_Caj'] / 1000 / df_rm['EMB'] * df_rm['Pedido']
        
    # Dinamica 1 
    # Crear la tabla pivote 
        pivot_table = pd.pivot_table(df_rm, 
                                    index=['Cliente', 'N° OC'], 
                                    values=['Pedido', 'Peso_Caj', 'Mto Saldo'], 
                                    aggfunc={'Pedido': 'sum', 'Peso_Caj': 'sum', 'Mto Saldo': 'sum'})

    # Convertir la tabla pivote a un DataFrame para facilitar la manipulación
        pivot_df = pivot_table.reset_index()

    # Calcular los totales por cliente y añadir una fila 'Total' para cada cliente
        totales_por_cliente = pivot_df.groupby('Cliente').sum().reset_index()
        totales_por_cliente['N° OC'] = 'Total'
                
    # Añadir las filas de totales al DataFrame original
        result_df = pd.concat([pivot_df, totales_por_cliente], ignore_index=True)
        
        
    #Eliminar todos las Oc's donde pedido sea 1 pero manteniendo la primera
        result_df_1=result_df[(result_df['Pedido'] == 1)]
        result_df_total =result_df[(result_df['N° OC']== 'Total')]
        result_df_first = result_df_1.drop_duplicates(subset='Cliente', keep= 'first')
        result_df_no1 =  result_df[(result_df['Pedido'] != 1) & (result_df['N° OC'] != 'Total')]
        result_df = pd.concat([result_df_first, result_df_no1, result_df_total]).sort_index()
    
    # Asegurarse de que las filas de totales están al final de cada cliente
        result_df['Cliente_order'] = result_df['N° OC'].apply(lambda x: 1 if x == 'Total' else 0)
        result_df = result_df.sort_values(by=['Cliente', 'Cliente_order', 'N° OC'], ascending=[True, True, True]).drop(columns=['Cliente_order'])
    
    # Se calcula el total general    
        total_general = result_df.drop(columns=['Cliente', 'N° OC']).sum().to_dict()
        total_general['Cliente'] = 'Total General'
        total_general['N° OC'] = 'Total'

    # Añadir la fila de total general al DataFrame
        result_df = result_df._append(total_general, ignore_index=True)        

    # Se definen los nombres del archivo y las hojas
        fecha_actual = datetime.now()
        fecha_actual_str = fecha_actual.strftime('%d-%m')
        nombre_archivo = f'Planning_RM {fecha_actual_str}.xlsx'
        ruta = os.path.join(directorio, nombre_archivo)
        hoja_bruto = 'data_en_bruto'
        pivot_hoja = 'tabla 1'
    # Se crea DataFrame para la funcion foto sin considerar filas de vev
        df_foto = result_df[~result_df['Cliente'].str.contains('vev', case=False) | (result_df['N° OC'] == 'Total')]
        df_foto = df_foto.reset_index(drop=True)
        fotoDinamica(df_foto, region= 'RM')
        
    # Se guardan en un excel con los nombres previamente definidios    
        with pd.ExcelWriter(ruta, engine='openpyxl') as writer:
            df_rm.to_excel(writer, sheet_name=hoja_bruto, index=False)
            df_foto.to_excel(writer, sheet_name=pivot_hoja, index=False)
            workbook = writer.book
            worksheet = workbook[pivot_hoja]

    # Definir el formato de celda para las filas que contienen 'Total'
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
    # Iterar sobre las filas del DataFrame con los totales y aplicar el estilo de relleno
            for row_idx in range(len(df_foto)):
                if df_foto.loc[row_idx, 'N° OC'] == 'Total':
                    for col_idx in range(len(df_foto.columns)):
                        cell = worksheet.cell(row=row_idx + 2, column=col_idx + 1)
                        cell.fill = fill
        # Dinamica 2
        # Se crean el orden de los columnas y el fortmato de estas para ordenar mas adelante en el codigo
            sheet_name = 'tabla 2'
            column_order =  ['Descripcion', 'Pedido', 'Va', 'Despacho', 'Stk-Disp', 'Stk-Res', 'Peso_Caj', 'EMB']
            oc_order = ['N° OC', 'N° Pedido', 'Articulo','Descripcion', 'Pedido', 'Va', 'Despacho', 'Stk-Disp', 'Stk-Res','Peso_Caj', 'EMB']
            columns_to_add = ['N° OC', 'N° Pedido', 'Articulo','Descripcion', 'Pedido', 'Va', 'Despacho', 'Stk-Disp', 'Stk-Res']
            
        # Se crean el orden de los columnas y el fortmato de estas para ordenar mas adelante en el codigo 
            if sheet_name not in writer.sheets:
                writer.book.create_sheet(title=sheet_name)
        
        # Recorre el dataframe considerando cada posicion como cada cliente unico, este los aisla en un DataFrame unico
            pivot_table_start_row = 0
            for i, cliente in enumerate(df_rm['Cliente'].unique()):
                df_cliente = df_rm[df_rm['Cliente'] == cliente]
        
        # Crea un DF para poder añadir el total general y se le da el formato previamente definido  
                suma_cliente = df_cliente.groupby('Cliente').agg({'Peso_Caj': 'sum', 'EMB': 'sum'}).reset_index()
                for col in columns_to_add:
                    suma_cliente[col] = ''
                suma_cliente = suma_cliente[oc_order]
                suma_cliente['N° OC'] = 'Total'
                suma_cliente['N° Pedido'] = 'General'
                suma_cliente['Articulo'] = ':'
        # Se crean 2 tablas pivot para el cliente y otra el total general de este
                pivot_table_cliente = pd.pivot_table(df_cliente,
                                                    index=['N° OC', 'N° Pedido', 'Articulo'],
                                                    values=['Descripcion', 'Pedido', 'Va', 'Despacho', 'Stk-Disp', 'Stk-Res', 'Peso_Caj', 'EMB'],
                                                    aggfunc='first')
                pivot_table_suma_cliente = pd.pivot_table(suma_cliente,
                                                    index=['N° OC', 'N° Pedido', 'Articulo'],
                                                    values=['Descripcion', 'Pedido', 'Va', 'Despacho', 'Stk-Disp', 'Stk-Res', 'Peso_Caj', 'EMB'],
                                                    aggfunc='first')
                
        # Se concatenan y se ordenan con el orden de las columnas buscados    
                pivot_table_cliente = pd.concat([pivot_table_cliente, pivot_table_suma_cliente])
                pivot_table_cliente = pivot_table_cliente[column_order]
        
        # Se guardan cada tabla unica por cliente en la hoja y se da formato.
                start_row = pivot_table_start_row
                sheet = writer.book[sheet_name]
                sheet.cell(row=start_row + 1, column=1).value = f'Cliente: {cliente}'
                pivot_table_cliente.to_excel(writer, sheet_name=sheet_name, startrow=start_row + 2, index=True)
                pivot_table_start_row = start_row + len(pivot_table_cliente) + 4
                
def casoRegiones(archivo_df):
    
    # Se recibe y se elimminan los datos que no se considentan
        df_rm = archivo_df
        columns_to_check = ['Sucursal', 'Nombre Sucursal', 'F12', 'Vendedor', 'Atencion', 'N° Factura', 
                   'Fec-Vcto', 'Fec-A-Partir-De', 'Fec-Imp.', 'Fec-Desp.', 'Ubicacion', 
                   'Precio_Lp', 'Precio_Bol', 'Dcto', 'Q-Ped', 'Q-Desp', 'Mto Pedido', 
                   'Mto Desp', 'Largo_Caj', 'Ancho_Caj', 'Alto_Caj', 'Stk_Otros', 'Stk_Lib', 
                   '1A1-STOCK', '1J1-JAULA', '1P1-PISO', '1P2-PISO', '3I3-RACK', '3I7-RACK', 
                   '6I2-RACK3', 'Comentarios Articulo']
        df_rm = df_rm[~((df_rm["Vendedor"] >30))]
        df_rm = df_rm[~((df_rm['Vendedor']<4))]
        columns_to_drop = [col for col in columns_to_check if col in df_rm.columns]
        df_rm = df_rm.drop(columns=columns_to_drop)
        
    # Se hacen cambios en el formato del excel recibido 
        df_rm = df_rm.rename(columns= {'Observacion': 'Transporte'})
        df_rm['Va'] = 'Si'
        df_rm['Despacho'] = formatearFecha()
        df_rm = df_rm[~((df_rm['Q-Saldo']==0))]
        df_rm = df_rm.rename(columns={'Q-Saldo': 'Pedido'})
        df_rm = df_rm.rename(columns={'Embaje': 'EMB'})
        
    # Se realizan calculos de volumen 
        df_rm['Peso_Caj'] = df_rm['Vol_Caj'] / 1000 / df_rm['EMB'] * df_rm['Pedido']
        
    # Dinamica 1 
    # Crear la tabla pivote 
        pivot_table = pd.pivot_table(df_rm, 
                                    index=['Cliente', 'N° Pedido'], 
                                    values=['Pedido', 'Peso_Caj', 'Mto Saldo', 'Va'], 
                                    aggfunc={'Pedido': 'sum', 'Peso_Caj': 'sum', 'Mto Saldo': 'sum'})

    # Convertir la tabla pivote a un DataFrame para facilitar la manipulación
        pivot_df = pivot_table.reset_index()

    # Calcular los totales por cliente y añadir una fila 'Total' para cada cliente
        totales_por_cliente = pivot_df.groupby('Cliente').sum().reset_index()
        totales_por_cliente['N° Pedido'] = 'Total'

    # Añadir las filas de totales al DataFrame original
        result_df = pd.concat([pivot_df, totales_por_cliente], ignore_index=True)
    
    # Asegurarse de que las filas de totales están al final de cada cliente
        result_df['Cliente_order'] = result_df['N° Pedido'].apply(lambda x: 1 if x == 'Total' else 0)
        result_df = result_df.sort_values(by=['Cliente', 'Cliente_order', 'N° Pedido'], ascending=[True, True, True]).drop(columns=['Cliente_order'])
        
    # Se calcula el total general    
        total_general = result_df.drop(columns=['Cliente', 'N° Pedido']).sum().to_dict()
        total_general['Cliente'] = 'Total General'
        total_general['N° Pedido'] = 'Total'

    # Añadir la fila de total general al DataFrame
        result_df = result_df._append(total_general, ignore_index=True)
    
    # Guardar el DataFrame original y la tabla dinámica con totales en un archivo Excel
        fecha_actual = datetime.now()
        fecha_actual_str = fecha_actual.strftime('%d-%m')
        nombre_archivo = f'Planning_Regiones {fecha_actual_str}.xlsx'
        ruta = os.path.join(directorio, nombre_archivo)
        hoja_bruto = 'data_en_bruto'
        pivot_hoja = 'tabla 1'
    # Se crea DataFrame para la funcion foto sin considerar filas de vev
        df_foto = result_df[~result_df['Cliente'].str.contains('vev', case=False) | (result_df['N° Pedido'] == 'Total')]
        df_foto = df_foto.reset_index(drop=True)
        fotoDinamica(df_foto, region= 'RG')
    # Se guardan los dataframe en hojas 
        with pd.ExcelWriter(ruta, engine='openpyxl') as writer:
            df_rm.to_excel(writer, sheet_name=hoja_bruto, index=False)
            df_foto.to_excel(writer, sheet_name=pivot_hoja, index=False)
            workbook = writer.book
            worksheet = workbook[pivot_hoja]

        # Definir el formato de celda para las filas que contienen 'Total'
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        # Iterar sobre las filas del DataFrame con los totales y aplicar el estilo de relleno
            for row_idx in range(len(df_foto)):
                if df_foto.loc[row_idx, 'N° Pedido'] == 'Total':
                    for col_idx in range(len(df_foto.columns)):
                        cell = worksheet.cell(row=row_idx + 2, column=col_idx + 1)
                        cell.fill = fill
        # Dinamica 2
        # Se crean el orden de los columnas y el fortmato de estas para ordenar mas adelante en el codigo
            sheet_name = 'tabla2 2'
            column_order =  ['Descripcion', 'Pedido', 'Va', 'Despacho', 'Stk-Disp', 'Stk-Res', 'Peso_Caj', 'EMB']
            oc_order = ['N° OC', 'N° Pedido', 'Articulo','Descripcion', 'Pedido', 'Va', 'Despacho', 'Stk-Disp', 'Stk-Res','Peso_Caj', 'EMB']
            columns_to_add = ['N° OC', 'N° Pedido', 'Articulo','Descripcion', 'Pedido', 'Va', 'Despacho', 'Stk-Disp', 'Stk-Res']
            
        #verficia la exsitencia de la hoja en el documento si no la crea
            if sheet_name not in writer.sheets:
                writer.book.create_sheet(title=sheet_name)
            pivot_table_start_row = 0
        
        # Recorre el dataframe considerando cada posicion como cada cliente unico, este los aisla en un DataFrame unico    
            for i, cliente in enumerate(df_rm['Cliente'].unique()):
                df_cliente = df_rm[df_rm['Cliente'] == cliente]
        
        # Crea un DF para poder añadir el total general y se le da el formato previamente definido        
                suma_cliente = df_cliente.groupby('Cliente').agg({'Peso_Caj': 'sum', 'EMB': 'sum'}).reset_index()
                for col in columns_to_add:
                    suma_cliente[col] = ''
                suma_cliente = suma_cliente[oc_order]
                suma_cliente['N° Pedido'] = 'Total'
                suma_cliente['Transporte'] = 'General'
                suma_cliente['Articulo'] = ':'
        
        # Se crean 2 tablas pivot para el cliente y otra el total general de este
                pivot_table_cliente = pd.pivot_table(df_cliente,
                                                    index=['N° Pedido', 'Transporte', 'Articulo'],
                                                    values=['Descripcion', 'Pedido', 'Va', 'Despacho', 'Stk-Disp', 'Stk-Res', 'Peso_Caj', 'EMB'],
                                                    aggfunc='first')
                pivot_table_suma_cliente = pd.pivot_table(suma_cliente,
                                                    index=['N° Pedido', 'Transporte', 'Articulo'],
                                                    values=['Descripcion', 'Pedido', 'Va', 'Despacho', 'Stk-Disp', 'Stk-Res', 'Peso_Caj', 'EMB'],
                                                    aggfunc='first')
                
        # Se concatenan y se ordenan con el orden de las columnas buscados       
                pivot_table_cliente = pd.concat([pivot_table_cliente, pivot_table_suma_cliente])
                pivot_table_cliente = pivot_table_cliente[column_order]
                
        # Se guardan cada tabla unica por cliente en la hoja y se da formato.        
                start_row = pivot_table_start_row
                sheet = writer.book[sheet_name]
                sheet.cell(row=start_row + 1, column=1).value = f'Cliente: {cliente}'
                pivot_table_cliente.to_excel(writer, sheet_name=sheet_name, startrow=start_row + 2, index=True)
                pivot_table_start_row = start_row + len(pivot_table_cliente) + 4
            
# Funcion para abrir el explorador de archivos
def seleccionar_archivo():
    archivo = filedialog.askopenfilename(filetypes=[("Archivos XLSX", "*.xlsx")])
    if archivo:
        abrir_ventana_proceso(archivo)

#Funcion que genera la ventana donde elegimos el tipo de planning
def abrir_ventana_proceso(archivo):
    root.withdraw()

    ventana_proceso = tk.Toplevel()
    ventana_proceso.title("Seleccionar Tipo de Planning")
    ventana_proceso.iconbitmap(ruta_icono)

    ancho_pantalla = ventana_proceso.winfo_screenwidth()
    alto_pantalla = ventana_proceso.winfo_screenheight()
    ancho_ventana = 400
    alto_ventana = 200
    x = (ancho_pantalla // 2) - (ancho_ventana // 2)
    y = (alto_pantalla // 2) - (alto_ventana // 2)
    ventana_proceso.geometry(f"{ancho_ventana}x{alto_ventana}+{x}+{y}")

    frame = tk.Frame(ventana_proceso)
    frame.pack(expand=True)

    planning_rm_btn = tk.Button(frame, text="Planning RM", command=lambda: procesar_archivo(ventana_proceso, archivo, "rm"))
    planning_rm_btn.pack(pady=10, padx=20)

    planning_regiones_btn = tk.Button(frame, text="Planning Regiones", command=lambda: procesar_archivo(ventana_proceso, archivo, "regiones"))
    planning_regiones_btn.pack(pady=10, padx=20)

# Funcion que desencadena la funcion de planning de acuerdo a la opcion seleccionada
def procesar_archivo(ventana_proceso, archivo, tipo_planning):
    try:
        archivo_df = pd.read_excel(archivo)

        if tipo_planning == "rm":
            casoRM(archivo_df)
        elif tipo_planning == "regiones":
            casoRegiones(archivo_df)

        ventana_proceso.destroy()
        abrir_ventana_final(tipo_planning)
    except pd.errors.ParserError as e:
        print(f"Error de parsing al leer el archivo XLSX: {e}")

#Funcion de la ventana que gestiona la llamada a la impresion de mail
def abrir_ventana_final(tipo_planning):
    ventana_final = tk.Toplevel()
    ventana_final.title("Finalización del Proceso")
    ventana_final.iconbitmap(ruta_icono)
    
    ancho_pantalla = ventana_final.winfo_screenwidth()
    alto_pantalla = ventana_final.winfo_screenheight()
    ancho_ventana = 400
    alto_ventana = 200
    x = (ancho_pantalla // 2) - (ancho_ventana // 2)
    y = (alto_pantalla // 2) - (alto_ventana // 2)
    ventana_final.geometry(f"{ancho_ventana}x{alto_ventana}+{x}+{y}")

    frame = tk.Frame(ventana_final)
    frame.pack(expand=True)

    enviar_email_btn = tk.Button(frame, text="Enviar Email", command=lambda: solicitar_credenciales(ventana_final, tipo_planning))
    enviar_email_btn.pack(pady=10, padx=20)
    cerrar_programa_btn = tk.Button(frame, text="Cerrar Programa", command=cerrar_programa)
    cerrar_programa_btn.pack(pady=10, padx=20)

#Funcion que gestiona la ventana de imgreso de credenciales
def solicitar_credenciales(ventana_final, tipo_planning):
    credenciales_ventana = tk.Toplevel(ventana_final)
    credenciales_ventana.title("Ingresar Credenciales")
    credenciales_ventana.iconbitmap(ruta_icono)

    ancho_pantalla = credenciales_ventana.winfo_screenwidth()
    alto_pantalla = credenciales_ventana.winfo_screenheight()
    ancho_ventana = 300
    alto_ventana = 200
    x = (ancho_pantalla // 2) - (ancho_ventana // 2)
    y = (alto_pantalla // 2) - (alto_ventana // 2)
    credenciales_ventana.geometry(f"{ancho_ventana}x{alto_ventana}+{x}+{y}")

    tk.Label(credenciales_ventana, text="Usuario:").pack(pady=5)
    usuario_entry = tk.Entry(credenciales_ventana)
    usuario_entry.pack(pady=5)

    tk.Label(credenciales_ventana, text="Contraseña:").pack(pady=5)
    contrasena_entry = tk.Entry(credenciales_ventana, show='*')
    contrasena_entry.pack(pady=5)

    botones_frame = tk.Frame(credenciales_ventana)
    botones_frame.pack(pady=10)

    def obtener_credenciales():
        usuario = usuario_entry.get()
        contrasena = contrasena_entry.get()
        if usuario and contrasena:
            credenciales_ventana.destroy()
            send_email(ventana_final, usuario, contrasena, tipo_planning)
        else:
            messagebox.showerror("Error", "Debe ingresar usuario y contraseña.")

    tk.Button(botones_frame, text="Enviar", command=obtener_credenciales).grid(row=0, column=0, padx=10)
    tk.Button(botones_frame, text="Cancelar", command=credenciales_ventana.destroy).grid(row=0, column=1, padx=10)
     
    
def send_email(ventana_final, usuario, contrasena, tipo_planning):

    #Definimos tipo de planning que ira en el asunto.
    if tipo_planning == 'rm':
        tipo_planning = 'Region Metropolitana'
    elif tipo_planning == 'regiones':
        tipo_planning = 'Regiones'
    subject = f'Planning {tipo_planning}'
    
    #Definimos el dia para buscar los archivos
    fecha_actual = datetime.now()
    fecha_actual_str = fecha_actual.strftime('%d-%m')
    
    #Definimos las rutas de los archivo de los destinatarios directos, con copia, la imagen y el excel
    #Destinatarios To y Cc
    #Imagen
    if tipo_planning == 'Region Metropolitana':
        image = fr'{ruta_RM} {fecha_actual_str}.png'
    elif tipo_planning == 'Regiones':
        image = fr'{ruta_RG} {fecha_actual_str}.png'
    #Excel
    if tipo_planning == 'Region Metropolitana':
        excel_path = fr'{ruta_RM} {fecha_actual_str}.xlsx'
    elif tipo_planning == 'Regiones':
        excel_path = fr'{ruta_RG} {fecha_actual_str}.xlsx'
    
    #Abrimos los archivos de los destinatorios directos y con copia para guardarlos en sus respectivas variables
    with open(archive_to, 'r') as archivo:
        to_emails = archivo.read().strip()
        to_emails = ast.literal_eval(to_emails)
    with open(archive_cc, 'r') as archivo:
        cc_emails = archivo.read().strip()
        cc_emails = ast.literal_eval(cc_emails)
    
    #definimos el cuerpo del mensaje y el html para incluir la imagen
    body = 'Estimados hago envio de planning del dia de hoy'
    html = f"""
    <html>
    <body>
        <p>{body}</p>
        <img src="cid:image1">
        <p>Saludos Cordiales</p>
    </body>
    </html>
    """
    #Iniciamos el mensaje y le damos los detalles para el mensaje
    msg = MIMEMultipart()
    msg['From'] = usuario
    msg['To'] = ', '.join(to_emails)
    msg['Cc'] = ', '.join(cc_emails)
    msg['Subject'] = subject
    msg.attach(MIMEText(html, 'html'))
    with open(image, 'rb') as img:
        mime_image = MIMEImage(img.read())
        mime_image.add_header('Content-ID', '<image1>')
        msg.attach(mime_image)
    with open(excel_path, 'rb') as attachment:
        part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header(
            'Content-Disposition',
            f'attachment; filename= Planning {tipo_planning} {fecha_actual_str}.xlsx',
        )
        msg.attach(part)
    all_recipients = to_emails + cc_emails
    
    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(usuario, contrasena)
        text = msg.as_string()
        server.sendmail(usuario, all_recipients, text)
        server.quit()
        print('exito')
        messagebox.showinfo("Éxito", "Correo enviado exitosamente.")
    except smtplib.SMTPAuthenticationError:
        print('Error')
        messagebox.showerror("Error", "Correo o contraseña invalida. Por favor reintente")
        return
        
    except Exception as e:
        print(f"Error al enviar el correo: {e}")
        messagebox.showerror("Error", f"Error innesperado {e}. Archivos guardados en = carpeta data")
        return
    
    ventana_final.destroy()
    root.deiconify()

def cerrar_programa():
    root.destroy()

root = tk.Tk()
root.title("Automatización Planning Bodega")
root.iconbitmap(ruta_icono) 

ancho_pantalla = root.winfo_screenwidth()
alto_pantalla = root.winfo_screenheight()
ancho_ventana = 400
alto_ventana = 200
x = (ancho_pantalla // 2) - (ancho_ventana // 2)
y = (alto_pantalla // 2) - (alto_ventana // 2)
root.geometry(f"{ancho_ventana}x{alto_ventana}+{x}+{y}")

frame = tk.Frame(root)
frame.pack(expand=True)

seleccionar_btn = tk.Button(frame, text="Seleccionar archivo XLSX", command=seleccionar_archivo)
seleccionar_btn.pack(pady=20, padx=20, anchor="center")

root.mainloop()